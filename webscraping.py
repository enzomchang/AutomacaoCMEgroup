""" AUTOMAÇÃO PARA PEGAR AS INFORMAÇÕES DO SITE CME GROUP E TRANSFORMAR EM PLANILHA """

""" Variações de preço em relação ao dia anterior, dados relevantes para traders e investidores no mercado de commodities agrícolas. 
Essas informações são essenciais para a análise do mercado e tomada de decisões de investimento. """

""" Para isso, foi disponibilizado uma planilha com as fórmulas necessárias do excel """

"""
Meu objetivo é pegar os seguintes dados do site:

* MONTH
* OPEN
* HIGH
* LOW
* LAST 
* CHANGE
* SETTLE
* EST. VOLUME
* PRIOR DAY OI

"""

"""
1o passo: Realizar os clickes com selenium

2o passo: Realizar o Webscraping da página

3o passo: Transformar os dados em planilha

4o passo: Realizar os tratamentos dos dados 

5o passo: Modificar o arquivo .env com o caminho necessário para salvar o arquivo para a variável CAMINHO_BASE

6o passo: Subscrever planilha FUTURES 06-08-2024 para a data atual e com os valores atualizados

"""

"""

### Mudanças necessárias para cada usuário: ( Estou sinalizando no código ) ###

- Modificar o arquivo .env de acordo com o e-mail que você quer que receba as promoções

- Agendar a tarefa de acordo com o que você quer que seja realizado o webscraping

"""


"""
Este script:
- Realiza web scraping para obter os dados necessários.
- Mescla os dados em um DataFrame.
- Salva os dados em uma planilha Excel.
- Realiza os tratamentos dos dados
- Atualiza a planilha "FUTURES 06-08-2024" com os dados obtidos, preservando a fórmula na coluna "Settlment" usando openpyxl.
"""

##############################
### Instalações Essenciais ###
##############################

# !pip install pandas
# !pip install numpy
# !pip install selenium
# !pip install webdriver-manager
# !pip install pywin32
# !pip install scrapy
# !pip install openpyxl
# !pip install schedule
# !pip install openpyxl

##############################
### Bibliotecas Essenciais ###
##############################

import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import time
from openpyxl import load_workbook
from datetime import datetime, timedelta
import numpy as np
import calendar
from openpyxl.styles import PatternFill
from dotenv import load_dotenv
import os

##############################
###### Funções Projeto #######
##############################

# Função para calcular o dia útil anterior
def dia_util_anterior(data):
    dia = data
    while True:
        dia -= timedelta(days=1)
        if dia.weekday() < 5:  # 0=segunda, 1=terça, ..., 4=sexta
            break
    return dia

# Função para configurar o navegador
def configurar_navegador():
    options = webdriver.ChromeOptions()
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, Gecko) Chrome/122.0.0.0 Safari/537.36")
    serviço = Service(ChromeDriverManager().install())
    navegador = webdriver.Chrome(service=serviço, options=options)
    return navegador

# Função para coletar os meses
def coletar_meses(navegador):
    meses = []
    try:
        tabela = WebDriverWait(navegador, 30).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="productTabData"]/div/div/div/div/div/div[2]/div/div/div/div/div/div[8]/div'))
        )
        linhas = tabela.find_elements(By.XPATH, './/tbody/tr')
        for linha in linhas:
            colunas = linha.find_elements(By.TAG_NAME, 'td')
            if colunas:
                mes = colunas[0].text
                if mes.strip():
                    meses.append(mes)
    except (NoSuchElementException, TimeoutException) as e:
        print(f"Erro ao encontrar a tabela: {e}")
    return pd.DataFrame(meses, columns=["MONTH"])

# Função para coletar os outros dados
def coletar_dados(navegador):
    dados = []
    try:
        tabela = WebDriverWait(navegador, 30).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="productTabData"]/div/div/div/div/div/div[2]/div/div/div/div/div/div[8]/div'))
        )
        linhas = tabela.find_elements(By.XPATH, './/tbody/tr')
        for linha in linhas:
            colunas = linha.find_elements(By.TAG_NAME, 'td')
            if colunas:
                linha_dados = [coluna.text for coluna in colunas[1:]]
                dados.append(linha_dados)
    except (NoSuchElementException, TimeoutException) as e:
        print(f"Erro ao encontrar a tabela: {e}")
    colunas = ["OPEN", "HIGH", "LOW", "LAST", "CHANGE", "SETTLE", "EST. VOLUME", "PRIOR DAY OI"]
    return pd.DataFrame(dados, columns=colunas)

""" Realizar o Webscraping """

# Função principal
def main():
    load_dotenv()  # Carrega as variáveis de ambiente do arquivo .env
    CAMINHO_BASE = os.getenv('CAMINHO_BASE')  # Obtém o caminho base do arquivo .env

    # Primeiro web scraping para coletar os meses
    navegador = configurar_navegador()
    navegador.get("https://www.cmegroup.com/markets/agriculture/oilseeds/soybean-meal.settlements.html")
    time.sleep(10)
    navegador.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(5)
    try:
        load_all_button = WebDriverWait(navegador, 30).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="productTabData"]/div/div/div/div/div/div[2]/div/div/div/div/div/div[9]/div[2]/button'))
        )
        navegador.execute_script("arguments[0].click();", load_all_button)
        time.sleep(10)
    except (NoSuchElementException, TimeoutException) as e:
        print(f"Erro ao clicar no botão 'Load All': {e}")
    
    df_meses = coletar_meses(navegador)
    navegador.quit()

    # Segundo web scraping para coletar os outros dados
    navegador = configurar_navegador()
    navegador.get("https://www.cmegroup.com/markets/agriculture/oilseeds/soybean-meal.settlements.html")
    time.sleep(10)
    navegador.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(5)
    try:
        load_all_button = WebDriverWait(navegador, 30).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="productTabData"]/div/div/div/div/div/div[2]/div/div/div/div/div/div[9]/div[2]/button'))
        )
        navegador.execute_script("arguments[0].click();", load_all_button)
        time.sleep(10)
    except (NoSuchElementException, TimeoutException) as e:
        print(f"Erro ao clicar no botão 'Load All': {e}")
    
    df_dados = coletar_dados(navegador)
    navegador.quit()

    ##############################
    ##### Programa Principal #####
    ##############################

    """ Tratamento dos dados """

    # Mesclar os DataFrames
    df_final = pd.concat([df_meses, df_dados], axis=1)

    # Inserir coluna em branco entre "PRIOR DAY OI" e "Operação"
    df_final.insert(df_final.columns.get_loc('PRIOR DAY OI') + 1, "", "")
    
    # Adicionar as colunas "Operação", "Ativo", "Série", e "SettIment" com valores até a linha 24
    df_final["Operação"] = ["FUTURE" if i < 23 else "" for i in range(len(df_final))]
    df_final["Ativo"] = ["SB" if i < 23 else "" for i in range(len(df_final))]
    df_final["Série"] = df_final['MONTH'].str.replace(" ", "")
    df_final["SettIment"] = df_final['SETTLE']
    
    # Reordenar as colunas para que "Operação" e "Ativo" estejam no lugar correto
    colunas_final = df_final.columns.tolist()
    df_final = df_final[colunas_final[:len(colunas_final) - 4] + colunas_final[-4:]]

    # Obter a data de hoje e calcular o dia anterior
    data_hoje = datetime.today()
    data_dia_anterior = data_hoje - timedelta(days=1)

    # Formatar a data para o nome do arquivo
    nome_arquivo = f"FUTURES {data_dia_anterior.strftime('%d-%m-%Y')}.xlsx"

    # Caminho do arquivo antigo
    caminho_antigo = os.path.join(CAMINHO_BASE, 'FUTURES 06-08-2024.xlsm') # Mudar de acordo com o seu caminho no arquivo .env
    
    # Carregar o arquivo existente
    wb = load_workbook(caminho_antigo)
    
    # Remover a planilha "SOJA" se ela existir
    if "SOJA" in wb.sheetnames:
        del wb["SOJA"]
    
    # Adicionar a nova planilha como "SOJA" na primeira posição
    new_sheet = wb.create_sheet(title="SOJA", index=0)

    # Definir as cores para o cabeçalho e os dados
    header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")  # Azul ênfase 5 mais escuro 25%
    data_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")    # Cinza ênfase 3, mais claro 80%

    # Limitar a estilização até a linha 23
    max_row = 23

    # Copiar os dados da nova planilha para o arquivo existente, incluindo o cabeçalho
    for col_idx, col_name in enumerate(df_final.columns, 1):
        cell = new_sheet.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill  # Aplicar a cor de fundo ao cabeçalho
    for r_idx, row in df_final.iterrows():
        if r_idx < max_row:
            for c_idx, value in enumerate(row):
                cell = new_sheet.cell(row=r_idx + 2, column=c_idx + 1, value=value)
                cell.fill = data_fill  # Aplicar a cor de fundo aos dados

    # Salvar o arquivo com o novo nome
    caminho_novo = os.path.join(CAMINHO_BASE, nome_arquivo)  # Mudar de acordo com o seu caminho no arquivo .env
    wb.save(caminho_novo)

    print(f"Arquivo salvo como {caminho_novo}")

# Executar a função principal
if __name__ == "__main__":
    main()

print("Fim do Programa")



